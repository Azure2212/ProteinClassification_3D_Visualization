"""Generate backend/data/pdb_part3.json from the professor's Excel workbook.

Part 3 for the **PDB** dataset is sourced DIRECTLY from
    sourceCode/visuallization/ExactNSimilarity229ProteinImages_ourSmoothing.xlsx
(229 real test images, top-50 only) — NOT from our own eval CSVs (those diverge
from the professor's reported numbers, so per user decision they are dropped for
PDB).  MAP Part 3 still comes from trained_results ExactN CSVs.

WHY a baked JSON:  the deployed site depends only on Flask (see requirements.txt);
openpyxl is a dev-only dependency.  This script is run offline once (with a Python
that has openpyxl) and commits the parsed numbers so the server needs no Excel.

KNOWN CORRUPTION handled here:  Excel auto-converted every "M/D" fraction string
whose first number is a valid month (1-12) into a real date, e.g. "5/11" ->
datetime(2026, 5, 11).  We recover the original fraction as f"{dt.month}/{dt.day}".
Recovery is validated: 9 of 12 rows reconstruct EXACTLY against the stated row
total.  The 3 baseline rows that don't (source-Excel sum typos) are flagged with
`mismatch:true` and both the computed and stated numerators are kept — nothing is
silently corrected or invented.

Run:  /path/to/python-with-openpyxl backend/gen_pdb_part3.py [path/to/xlsx]
"""
import os
import sys
import re
import json
import datetime

import openpyxl

# Canonical real-protein order (matches config.REAL_PROTEINS / MAP display).
CANON = ["7UZM", "8DNM", "8DNO", "8DNP", "8DNS", "8DNU",
         "8EM2", "8EMS", "8EMT", "8ENE", "8EOJ", "8EOR"]

# 12 protein data columns E..P in every block.
COLS = "EFGHIJKLMNOP"

# (sheet, model, input_shape, group, header_row, [(variant_label, data_row), ...])
BLOCKS = [
    ("Sheet3", "Resnet152", "224x224", "no smoothing", 16,
     [("filter=6", 17), ("filter=12", 18), ("filter=14", 19)]),
    ("Sheet3", "Swinv2b", "256x256", "no smoothing", 25,
     [("filter=6", 26), ("filter=12", 27), ("filter=14", 28)]),
    ("Sheet2", "Resnet152", "224x224", "smoothing (ε=0.2)", 16,
     [("without smoothing", 17), ("hard smoothing", 18), ("our smoothing", 19)]),
    ("Sheet2", "Swinv2b", "256x256", "smoothing (ε=0.2)", 28,
     [("without smoothing", 29), ("hard smoothing", 30), ("our smoothing", 31)]),
]


def recover(v):
    """Cell value -> (correct, total, raw, was_date) or None if not a fraction."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        # Excel turned "M/D" into a date; recover the fraction.
        return (v.month, v.day, f"{v.month}/{v.day}", True)
    m = re.match(r"^\s*(\d+)\s*/\s*(\d+)", str(v))
    if m:
        return (int(m.group(1)), int(m.group(2)), f"{m.group(1)}/{m.group(2)}", False)
    return None


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    default_xlsx = os.path.normpath(os.path.join(
        here, "..", "..", "sourceCode", "visuallization",
        "ExactNSimilarity229ProteinImages_ourSmoothing.xlsx"))
    xlsx = sys.argv[1] if len(sys.argv) > 1 else default_xlsx
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    models = {}
    protein_totals = {}
    for sheet, model, shape, group, hdr, rows in BLOCKS:
        ws = wb[sheet]
        header = [ws[f"{c}{hdr}"].value for c in COLS]
        m = models.setdefault(model, {"input": shape, "columns": []})
        for label, r in rows:
            stated = ws[f"Q{r}"].value
            sm = re.match(r"\s*(\d+)\s*/\s*(\d+)", str(stated) if stated else "")
            stated_correct = int(sm.group(1)) if sm else None
            cells = {}
            comp_correct = comp_total = 0
            recovered_cells = []
            for c, prot in zip(COLS, header):
                rec = recover(ws[f"{c}{r}"].value)
                if rec is None:
                    cells[prot] = None
                    continue
                cor, tot, raw, was_date = rec
                cells[prot] = {"correct": cor, "total": tot}
                comp_correct += cor
                comp_total += tot
                if was_date:
                    recovered_cells.append(prot)
                # track per-protein denominator (consistency check below)
                protein_totals.setdefault(prot, tot)
            mismatch = (stated_correct is not None and stated_correct != comp_correct)
            m["columns"].append({
                "key": f"{sheet}:{r}",
                "label": label,
                "group": group,
                "sheet": sheet,
                "cells": cells,
                "computed_correct": comp_correct,
                "computed_total": comp_total,
                "stated": str(stated) if stated is not None else None,
                "stated_correct": stated_correct,
                "mismatch": mismatch,
                "recovered_cells": recovered_cells,
            })

    proteins = [p for p in CANON if p in protein_totals]
    proteins += [p for p in protein_totals if p not in proteins]

    out = {
        "source": "sourceCode/visuallization/ExactNSimilarity229ProteinImages_ourSmoothing.xlsx",
        "note": ("PDB Part 3 = EMDB_ExactNSimilarity 229 (Excel). 229 real test "
                 "images, top-50 only. Model input: Resnet152 224x224, "
                 "Swinv2b 256x256. Sheet3 = per-filter (no smoothing); "
                 "Sheet2 = smoothing comparison (ε=0.2)."),
        "topk": 50,
        "total_images": sum(protein_totals.values()),
        "proteins": proteins,
        "protein_totals": {p: protein_totals[p] for p in proteins},
        "models": models,
    }

    outdir = os.path.join(here, "data")
    os.makedirs(outdir, exist_ok=True)
    outpath = os.path.join(outdir, "pdb_part3.json")
    with open(outpath, "w") as f:
        json.dump(out, f, indent=2)

    # console report
    print(f"source : {xlsx}")
    print(f"wrote  : {outpath}")
    print(f"total  : {out['total_images']} images, "
          f"{len(proteins)} proteins, {len(models)} models")
    for model, mv in models.items():
        print(f"  {model} ({mv['input']}): "
              f"{len(mv['columns'])} columns")
        for c in mv["columns"]:
            flag = "  <-- MISMATCH" if c["mismatch"] else ""
            print(f"    {c['group']:20s} {c['label']:20s} "
                  f"computed={c['computed_correct']}/{c['computed_total']} "
                  f"stated={c['stated']}{flag}")


if __name__ == "__main__":
    main()
